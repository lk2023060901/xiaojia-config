import openpyxl

def cleanup_beans(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # 查找并删除所有的 CompetitionConfig 定义，然后重新写一个唯一的
    rows_to_delete = []
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=2).value == "CompetitionConfig":
            # 找到起始行，向下找直到遇到空行
            curr = row
            while curr <= sheet.max_row:
                rows_to_delete.append(curr)
                if sheet.cell(row=curr + 1, column=10).value is None and \
                   (curr + 1 > sheet.max_row or sheet.cell(row=curr + 1, column=2).value is not None):
                    break
                curr += 1
    
    # 从后往前删，避免索引混乱
    for r in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(r)
        
    # 重新在末尾写一个干净的
    start_row = sheet.max_row + 2
    sheet.cell(row=start_row, column=2).value = "CompetitionConfig"
    sheet.cell(row=start_row, column=7).value = "竞赛专属配置"
    sheet.cell(row=start_row, column=9).value = "c;s"
    
    fields = [
        ("max_competitor_count", "int", "可以参与竞赛的人数上限"),
        ("prize_random_count", "int", "随机抽取奖励数量"),
        ("prize_fixed_count", "int", "固定展示奖励数量"),
        ("min_pk_score", "int", "抢占擂台最低分数要求"),
        ("min_extra_energy", "int", "额外助力最低能量限制"),
        ("pk_idle_timeout", "int", "无人抢占超时时长(秒)"),
        ("settle_close_secs", "int", "结算自动关闭时长(秒)")
    ]
    for i, (name, dtype, comment) in enumerate(fields):
        row = start_row + i
        sheet.cell(row=row, column=10).value = name
        sheet.cell(row=row, column=12).value = dtype
        sheet.cell(row=row, column=13).value = "c;s"
        sheet_beans = sheet.cell(row=row, column=14).value = comment

    wb.save(file_path)

if __name__ == "__main__":
    cleanup_beans("xiaojia-config/datas/__beans__.xlsx")
